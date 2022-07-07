from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

work_book = Workbook()
work_sheet = work_book.active
work_sheet_title = "Grades"

headings = ['Name'] + list(data['Joe'].keys())
work_sheet.append(headings)

for person in data:
    grades = list(data[person].values())
    work_sheet.append([person] + grades)

for person in data:
	grades = list(data[person].values())
	work_sheet.append([person] + grades)

for col in range(2, len(data['Joe']) + 2):
	char = get_column_letter(col)
	work_sheet[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
	work_sheet[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

work_book.save("Grades.xlsx") 